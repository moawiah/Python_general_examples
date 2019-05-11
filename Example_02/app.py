import openpyxl as xl
from openpyxl.chart import  BarChart, Reference

#Sheet Specifications
price_column = 3
row_offset = 2
sale_percentage = 0.1

def data_load(filename):
    try:
        workbook = xl.load_workbook(filename)
        sheet = workbook['Sheet1']
    except IOError:
        print("Cannot read from %s, please try again!"%(filename))
        exit(1)


    return workbook, sheet

#Perform the needed price modification
def modify_price(workbook, sheet, modified_filename):
    for row in range(row_offset, sheet.max_row + 1):
        cell = sheet.cell(row, price_column)
        modified_price = cell.value * (1.0 - sale_percentage)
        modified_price_cell = sheet.cell(row, price_column + 1)
        modified_price_cell.value = modified_price

        workbook.save(modified_filename)

def draw_chart(workbook, sheet, modified_filename):
    #Gettig the values to be used in a chart
    values = Reference(sheet,
                       min_row=row_offset,
                       max_row=sheet.max_row,
                       min_col=price_column + 1,
                       max_col=price_column + 1)

    modified_price_chart = BarChart()
    modified_price_chart.add_data(values)
    sheet.add_chart(modified_price_chart)

    workbook.save(modified_filename)

# Example code to test the implementation
wb, data_sheet = data_load('transactions.xlsx')
# calculate the discount and create a new column of modified prices
modify_price(workbook=wb,sheet=data_sheet,modified_filename='modified_transactions.xlsx')
# Draw a chart of the calculated data
draw_chart(workbook=wb,sheet=data_sheet,modified_filename='modified_transactions.xlsx')
